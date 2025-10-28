from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from pathlib import Path
import datetime as dt
from decimal import Decimal
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    raise CommandError("Falta openpyxl. Instala con: pip install openpyxl")

from cuentahabientes.models import Cuentahabiente
from colonia.models import Colonia
from servicio.models import Servicio
from pagos.models import Pago
from descuento.models import Descuento
from cobrador.models import Cobrador


COLUMNAS = {
    "numero_contrato": {"Numero_contrato", "Contrato", "NumContrato", "NumeroContrato"},
    "nombres": {"Nombres", "Nombre"},
    "ap": {"Apellido paterno", "AP", "Apellido_paterno"},
    "am": {"Apellido materno", "AM", "Apellido_materno"},
    "calle": {"Calle"},
    "numero": {"Numero", "No", "Num"},
    "telefono": {"Telefono", "Teléfono"},
    "colonia": {"Colonia"},
    "mes": {"Mes"},
    "anio": {"Año", "Anio"},
    "saldo_pendiente": {"Saldo_pendiente", "Saldo", "SaldoPendiente"},
    "monto_recibido": {"Monto_recibido", "Pago", "Monto"},
    "monto_descuento": {"Monto_descuento", "Descuento"},
}


def _pick(row_values, headers, key, default=None, cast=None):
    posibles = COLUMNAS.get(key, {key})
    for i, h in enumerate(headers):
        if h in posibles:
            val = row_values[i] if i < len(row_values) else None
            if val is None or val == "":
                return default
            if cast:
                try:
                    return cast(val)
                except Exception:
                    return default
            return val
    return default


def mes_a_num(mes):
    """Convierte nombre de mes a número"""
    if mes is None:
        return 1
    mes = str(mes).strip().lower()
    mapa = {
        "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
        "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
        "septiembre": 9, "setiembre": 9, "octubre": 10,
        "noviembre": 11, "diciembre": 12,
    }
    return mapa.get(mes, 1)


def calcular_estatus(pagos, saldo_final, fecha_actual=None):
    """Calcula el estatus del cuentahabiente basándose en sus pagos"""
    if fecha_actual is None:
        fecha_actual = dt.date.today()
    
    # Si no tiene saldo pendiente, está pagado
    if saldo_final == 0:
        return "Pagado"
    
    # Contar cuántos meses únicos ha pagado
    meses_pagados = set()
    for pago in pagos:
        if pago["monto_recibido"] > 0 or pago["monto_descuento"] > 0:
            mes_num = mes_a_num(pago["mes"])
            anio = pago["anio"]
            meses_pagados.add((anio, mes_num))
    
    total_meses_pagados = len(meses_pagados)
    
    # Si no ha pagado nada o solo 1-2 meses
    if total_meses_pagados <= 2:
        return "Adeudo"
    
    # Verificar si está al corriente (pagado hasta mes actual o adelantado)
    mes_actual = fecha_actual.month
    anio_actual = fecha_actual.year
    
    # Buscar el último mes pagado
    if meses_pagados:
        meses_ordenados = sorted(meses_pagados, key=lambda x: (x[0], x[1]))
        ultimo_anio, ultimo_mes = meses_ordenados[-1]
        
        # Si pagó el mes actual o está adelantado
        if (ultimo_anio > anio_actual or 
            (ultimo_anio == anio_actual and ultimo_mes >= mes_actual)):
            return "Corriente"
    
    # Si pagó entre 3 y 6 meses (pero no está al corriente)
    if 3 <= total_meses_pagados <= 6:
        return "Rezagado"
    
    # Si pagó más de 6 meses pero no está al corriente
    if total_meses_pagados > 6:
        return "Rezagado"
    
    return "Adeudo"


class Command(BaseCommand):
    help = "Importa Cuentahabientes desde Excel con múltiples pagos por persona."

    def add_arguments(self, parser):
        parser.add_argument("ruta_excel", type=str, help="Ruta del archivo .xlsx")
        parser.add_argument("--hoja", type=str, default=None,
                            help="Nombre de la hoja. Si no se indica, se usa la primera.")
        parser.add_argument("--fila-header", type=int, default=1,
                            help="Número de fila donde están los encabezados (default: 1).")
        parser.add_argument("--servicio", type=str, required=True,
                            help="Nombre del Servicio a asignar.")
        parser.add_argument("--cobrador", type=str, required=True,
                            help="Usuario del cobrador para los pagos.")
        parser.add_argument("--crear-pagos", action="store_true",
                            help="Crea registros de Pago por cada fila.")
        parser.add_argument("--generar-contratos", action="store_true",
                            help="Genera números de contrato automáticamente.")
        parser.add_argument("--base-contrato", type=int, default=10000,
                            help="Número base para generar contratos (default: 10000).")
        parser.add_argument("--dry-run", action="store_true",
                            help="Simula sin escribir cambios.")

    def handle(self, *args, **opts):
        ruta = Path(opts["ruta_excel"]).expanduser()
        if not ruta.exists():
            raise CommandError(f"No existe el archivo: {ruta}")

        hoja = opts["hoja"]
        fila_header = opts["fila_header"]
        dry = opts["dry_run"]
        crear_pagos = opts["crear_pagos"]
        generar_contratos = opts["generar_contratos"]
        base_contrato = opts["base_contrato"]
        servicio_nombre = opts["servicio"]
        cobrador_usuario = opts["cobrador"]

        # Validaciones previas
        try:
            servicio = Servicio.objects.get(nombre__iexact=servicio_nombre)
        except Servicio.DoesNotExist:
            raise CommandError(f"Servicio '{servicio_nombre}' no encontrado.")

        try:
            cobrador = Cobrador.objects.get(usuario__iexact=cobrador_usuario)
        except Cobrador.DoesNotExist:
            raise CommandError(f"Cobrador '{cobrador_usuario}' no encontrado.")

        desc_pp = Descuento.objects.filter(nombre_descuento__iexact="Promoción Anual").first()
        desc_inapam = Descuento.objects.filter(nombre_descuento__iexact="INAPAM").first()

        # Abre el Excel
        wb = openpyxl.load_workbook(ruta, data_only=True)
        ws = wb[hoja] if hoja else wb.worksheets[0]
        headers = [str(c or "").strip() for c in next(ws.iter_rows(min_row=fila_header, max_row=fila_header, values_only=True))]
        
        self.stdout.write(self.style.WARNING(f"Headers detectados: {headers}"))

        # === PASO 1: AGRUPAR DATOS POR CUENTAHABIENTE ===
        cuentahabientes_data = {}
        filas_procesadas = 0
        filas_saltadas = 0
        
        for row_values in ws.iter_rows(min_row=fila_header + 1, values_only=True):
            filas_procesadas += 1
            try:
                numero_contrato = _pick(row_values, headers, "numero_contrato", cast=int)
                nombres = (_pick(row_values, headers, "nombres", default="") or "").strip()
                ap = (_pick(row_values, headers, "ap", default="") or "").strip()
                am = (_pick(row_values, headers, "am", default="") or "").strip()
                nombre_colonia = (_pick(row_values, headers, "colonia", default="") or "").strip()

                if filas_procesadas == 1:
                    self.stdout.write(self.style.WARNING(
                        f"Primera fila - Contrato: {numero_contrato}, "
                        f"Nombre: {nombres}, AP: {ap}, Colonia: {nombre_colonia}"
                    ))

                if not nombres and not ap and not numero_contrato:
                    filas_saltadas += 1
                    continue
                if not nombre_colonia:
                    filas_saltadas += 1
                    continue

                if numero_contrato:
                    clave = (numero_contrato, None, None, None, nombre_colonia)
                else:
                    clave = (None, nombres.lower(), ap.lower(), am.lower(), nombre_colonia)

                if clave not in cuentahabientes_data:
                    cuentahabientes_data[clave] = {
                        "numero_contrato": numero_contrato,
                        "nombres": nombres,
                        "ap": ap,
                        "am": am,
                        "calle": (_pick(row_values, headers, "calle", default="S/N") or "").strip(),
                        "numero": _pick(row_values, headers, "numero", default=0, cast=int) or 0,
                        "telefono": (_pick(row_values, headers, "telefono", default="S/N") or "").strip(),
                        "colonia": nombre_colonia,
                        "pagos": []
                    }

                monto_recibido = _pick(row_values, headers, "monto_recibido", default=0)
                try:
                    monto_recibido = int(monto_recibido or 0)
                except:
                    monto_recibido = 0

                monto_descuento = _pick(row_values, headers, "monto_descuento", default=0)
                try:
                    monto_descuento = int(monto_descuento or 0)
                except:
                    monto_descuento = 0

                if monto_recibido > 0 or monto_descuento > 0:
                    mes = _pick(row_values, headers, "mes", default="Enero")
                    anio = _pick(row_values, headers, "anio", default=dt.date.today().year, cast=int)
                    saldo_pendiente = _pick(row_values, headers, "saldo_pendiente", default=None)
                    
                    cuentahabientes_data[clave]["pagos"].append({
                        "mes": mes,
                        "anio": anio,
                        "monto_recibido": monto_recibido,
                        "monto_descuento": monto_descuento,
                        "saldo_pendiente": saldo_pendiente
                    })

            except Exception as e:
                self.stderr.write(self.style.WARNING(f"[Lectura fila {filas_procesadas}] Error: {e}"))
                continue

        self.stdout.write(self.style.WARNING(
            f"Filas procesadas: {filas_procesadas}, Saltadas: {filas_saltadas}, "
            f"Cuentahabientes únicos encontrados: {len(cuentahabientes_data)}"
        ))

        # === PASO 2: PROCESAR CUENTAHABIENTES ÚNICOS ===
        creados = 0
        actualizados = 0
        pagos_creados = 0
        errores = 0
        contratos_generados = 0

        if generar_contratos:
            ultimo_contrato = Cuentahabiente.objects.filter(
                numero_contrato__isnull=False
            ).order_by('-numero_contrato').values_list('numero_contrato', flat=True).first()
            
            proximo_contrato = max(base_contrato, (ultimo_contrato or 0) + 1) if ultimo_contrato else base_contrato
        else:
            proximo_contrato = None

        with transaction.atomic():
            try:
                for clave, datos in cuentahabientes_data.items():
                    try:
                        numero_contrato = datos["numero_contrato"]
                        nombre_colonia = datos["colonia"]

                        try:
                            colonia = Colonia.objects.get(nombre_colonia__iexact=nombre_colonia)
                        except Colonia.DoesNotExist:
                            self.stderr.write(self.style.ERROR(f"Colonia '{nombre_colonia}' no existe."))
                            errores += 1
                            continue

                        # Calcular saldo final
                        saldo_final = None
                        
                        if datos["pagos"]:
                            pagos_ordenados = sorted(
                                datos["pagos"], 
                                key=lambda p: (p["anio"], mes_a_num(p["mes"]))
                            )
                            
                            ultimo_pago = pagos_ordenados[-1]
                            
                            if ultimo_pago["saldo_pendiente"] is not None:
                                try:
                                    saldo_antes = int(ultimo_pago["saldo_pendiente"])
                                    pago_aplicado = ultimo_pago["monto_recibido"] + ultimo_pago["monto_descuento"]
                                    saldo_final = max(0, saldo_antes - pago_aplicado)
                                except:
                                    saldo_final = 0
                            else:
                                saldo_final = 0
                        else:
                            saldo_final = int(Decimal(servicio.costo))
                        
                        # Calcular estatus
                        estatus = calcular_estatus(datos["pagos"], saldo_final)

                        # Crear o actualizar cuentahabiente
                        if numero_contrato:
                            ch, created = Cuentahabiente.objects.update_or_create(
                                numero_contrato=numero_contrato,
                                defaults={
                                    "nombres": datos["nombres"],
                                    "ap": datos["ap"],
                                    "am": datos["am"],
                                    "calle": datos["calle"],
                                    "numero": datos["numero"],
                                    "telefono": datos["telefono"],
                                    "colonia": colonia,
                                    "servicio": servicio,
                                    "saldo_pendiente": saldo_final,
                                    "deuda": estatus,
                                }
                            )
                        else:
                            if generar_contratos:
                                while Cuentahabiente.objects.filter(numero_contrato=proximo_contrato).exists():
                                    proximo_contrato += 1
                                
                                ch, created = Cuentahabiente.objects.update_or_create(
                                    numero_contrato=proximo_contrato,
                                    defaults={
                                        "nombres": datos["nombres"],
                                        "ap": datos["ap"],
                                        "am": datos["am"],
                                        "calle": datos["calle"],
                                        "numero": datos["numero"],
                                        "telefono": datos["telefono"],
                                        "colonia": colonia,
                                        "servicio": servicio,
                                        "saldo_pendiente": saldo_final,
                                        "deuda": estatus,
                                    }
                                )
                                contratos_generados += 1
                                proximo_contrato += 1
                            else:
                                ch = Cuentahabiente.objects.filter(
                                    nombres__iexact=datos["nombres"],
                                    ap__iexact=datos["ap"],
                                    am__iexact=datos["am"],
                                    colonia=colonia,
                                    numero_contrato__isnull=True
                                ).first()
                                
                                if ch:
                                    ch.calle = datos["calle"]
                                    ch.numero = datos["numero"]
                                    ch.telefono = datos["telefono"]
                                    ch.servicio = servicio
                                    ch.saldo_pendiente = saldo_final
                                    ch.deuda = estatus
                                    ch.save()
                                    created = False
                                else:
                                    ch = Cuentahabiente.objects.create(
                                        numero_contrato=None,
                                        nombres=datos["nombres"],
                                        ap=datos["ap"],
                                        am=datos["am"],
                                        calle=datos["calle"],
                                        numero=datos["numero"],
                                        telefono=datos["telefono"],
                                        colonia=colonia,
                                        servicio=servicio,
                                        saldo_pendiente=saldo_final,
                                        deuda=estatus,
                                    )
                                    created = True

                        if created:
                            creados += 1
                        else:
                            actualizados += 1

                        # Crear pagos
                        if crear_pagos:
                            for pago_data in datos["pagos"]:
                                try:
                                    fecha_pago = dt.date(
                                        int(pago_data["anio"]), 
                                        mes_a_num(pago_data["mes"]), 
                                        15
                                    )
                                except:
                                    fecha_pago = dt.date.today()

                                descuento = None
                                monto_desc = pago_data["monto_descuento"]
                                if monto_desc == 60 and desc_pp:
                                    descuento = desc_pp
                                elif monto_desc in (300, 360) and desc_inapam:
                                    descuento = desc_inapam

                                mes_texto = str(pago_data["mes"]).strip().capitalize()

                                Pago.objects.create(
                                    descuento=descuento,
                                    cobrador=cobrador,
                                    cuentahabiente=ch,
                                    fecha_pago=fecha_pago,
                                    monto_recibido=pago_data["monto_recibido"],
                                    monto_descuento=pago_data["monto_descuento"],
                                    mes=mes_texto,
                                    anio=int(pago_data["anio"]),
                                )
                                pagos_creados += 1

                    except Exception as e:
                        errores += 1
                        self.stderr.write(self.style.WARNING(f"[Procesamiento] Error: {e}"))
                        continue

                if dry:
                    transaction.set_rollback(True)

            except Exception as e:
                transaction.set_rollback(True)
                raise

        msg = f"OK. Cuentahabientes únicos: {len(cuentahabientes_data)} | "
        msg += f"Creados: {creados}, Actualizados: {actualizados} | "
        msg += f"Pagos: {pagos_creados}"
        if generar_contratos:
            msg += f" | Contratos generados: {contratos_generados}"
        msg += f" | Errores: {errores}"
        
        self.stdout.write(self.style.SUCCESS(msg))